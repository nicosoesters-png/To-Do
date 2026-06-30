import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "To-Do"

DARK_BLUE = "1F3864"
RED = "C00000"
ORANGE = "ED7D31"
YELLOW = "FFC000"
BLUE = "2E75B6"
GREY = "808080"

# ---------- Title ----------
ws["A1"] = "To-Do List"
ws["A1"].font = Font(name="Calibri", size=28, bold=True, color=DARK_BLUE)
ws.row_dimensions[1].height = 40

# ---------- Tasks table (columns G:K) ----------
headers = ["Task", "Priority", "Due", "Done", "Area"]
header_row = 2
for i, h in enumerate(headers):
    col = get_column_letter(7 + i)  # G=7
    ws[f"{col}{header_row}"] = h

today = datetime.date(2026, 6, 30)

example_tasks = [
    ("Finish proposal", "High", datetime.date(2026, 6, 30), False),
    ("Pay invoice", "Medium", datetime.date(2026, 6, 28), False),
    ("Call client", "Medium", datetime.date(2026, 7, 1), False),
    ("Team meeting prep", "Low", datetime.date(2026, 7, 3), False),
    ("Submit report", "High", datetime.date(2026, 7, 10), False),
    ("Clean inbox", "Low", None, False),
]

first_data_row = header_row + 1
last_data_row = first_data_row + len(example_tasks) - 1

for r, (task, prio, due, done) in enumerate(example_tasks, start=first_data_row):
    ws[f"G{r}"] = task
    ws[f"H{r}"] = prio
    if due is not None:
        c = ws[f"I{r}"]
        c.value = due
        c.number_format = "DD.MM.YYYY"
    ws[f"J{r}"] = done
    ws[f"K{r}"] = (
        f'=LET(d,[@Due],done,[@Done],IF(done=TRUE,"Done",'
        f'IF(d="","General",IF(d<=TODAY(),"Today",'
        f'IF(d=TODAY()+1,"Tomorrow",'
        f'IF(d<=TODAY()-WEEKDAY(TODAY(),2)+7,"This Week","Next Week"))))))'
    )

table_ref = f"G{header_row}:K{last_data_row}"
tab = Table(displayName="Tasks", ref=table_ref)
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
ws.add_table(tab)

# Priority dropdown
dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws.add_data_validation(dv_priority)
dv_priority.add(f"H{first_data_row}:H{last_data_row + 20}")

# Done as TRUE/FALSE boolean dropdown (select column + Insert > Checkbox in Excel 365
# turns these into native clickable checkboxes)
dv_done = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv_done)
dv_done.add(f"J{first_data_row}:J{last_data_row + 20}")
for r in range(first_data_row, last_data_row + 1):
    ws[f"J{r}"].alignment = Alignment(horizontal="center")

# Hide helper Area column
ws.column_dimensions["K"].hidden = True

# Conditional formatting: Done = TRUE -> grey + strikethrough across the table row
done_font = Font(strike=True, color="A6A6A6")
grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
cf_range = f"G{first_data_row}:J{last_data_row + 20}"
ws.conditional_formatting.add(
    cf_range,
    FormulaRule(formula=[f"$J{first_data_row}=TRUE"], font=done_font, fill=grey_fill, stopIfTrue=False),
)

# ---------- Overview boxes ----------
def style_header(cell, color, font_color="FFFFFF"):
    cell.font = Font(bold=True, size=13, color=font_color)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def filter_formula(area_name):
    return (
        f'=FILTER(Tasks[Task]&IF(Tasks[Due]="",""," ("&TEXT(Tasks[Due],"DD.MM.")&")"),'
        f'Tasks[Area]="{area_name}","— none —")'
    )

# Today: column A, header row 3
ws.merge_cells("A3:B3")
ws["A3"] = "Today"
style_header(ws["A3"], RED)
ws["A4"] = filter_formula("Today")

# Tomorrow: column D (gap C), header row 3
ws.merge_cells("D3:E3")
ws["D3"] = "Tomorrow"
style_header(ws["D3"], ORANGE)
ws["D4"] = filter_formula("Tomorrow")

# This Week: row 14, columns A:E
ws.merge_cells("A14:E14")
ws["A14"] = "This Week"
style_header(ws["A14"], YELLOW, font_color="000000")
ws["A15"] = filter_formula("This Week")

# Next Week: row 24, columns A:E
ws.merge_cells("A24:E24")
ws["A24"] = "Next Week"
style_header(ws["A24"], BLUE)
ws["A25"] = filter_formula("Next Week")

# General: row 34, columns A:E
ws.merge_cells("A34:E34")
ws["A34"] = "General"
style_header(ws["A34"], GREY)
ws["A35"] = filter_formula("General")

for cell in ["A4", "D4", "A15", "A25", "A35"]:
    ws[cell].alignment = Alignment(wrap_text=False, vertical="top")

# ---------- Column widths ----------
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 3
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 28
ws.column_dimensions["F"].width = 4
ws.column_dimensions["G"].width = 32
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 14
ws.column_dimensions["J"].width = 9
ws.column_dimensions["K"].width = 12

ws.sheet_view.showGridLines = False

wb.save("To-Do-List.xlsx")
print("saved")
